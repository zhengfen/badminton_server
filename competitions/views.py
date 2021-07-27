from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets
from competitions.models import Club, ClubResponsable, Contact, Structure, Team, User
from competitions.serializers import ClubListSerializer, ClubShowSerializer, StructureSerializer, TeamSerializer, ClubResponsableSerializer
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import load_workbook
import os
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
import logging
# Get an instance of a logger
logger = logging.getLogger(__name__)

# Create your views here.
class ClubViewSet(viewsets.ModelViewSet): 
    perms_map = {
        'get': '*', 
        'post': 'club_create', 
        'put': 'club_update', 
        'delete': 'club_delete',
    }

    queryset = Club.objects.all()
    serializer_class = ClubListSerializer
    # filterset_class
    pagination_class = None
    search_fields = ['name']
    ordering_fields = ['name']

    def list(self, request):
        clubs = Club.objects.all()
        clubs_s = ClubListSerializer(clubs, many=True)

        structures = Structure.objects.all()
        structures_s = StructureSerializer(structures, many=True)
        return Response({
            'clubs':clubs_s.data, 
            'structures':structures_s.data
        })

    def retrieve(self, request, pk=None):
        club = get_object_or_404(Club, pk=pk)
        club_s = ClubShowSerializer(club)
        club_responsables = club.responsables;         
        club_responsables_s = ClubResponsableSerializer(club_responsables, many=True)
        return Response({
            'club': club_s.data, 
            'responsables': club_responsables_s.data
        })
    

class TeamViewSet(viewsets.ModelViewSet):
    queryset = Team.objects.all()
    serializer_class = TeamSerializer


def parse_responsables(request): 
    filename = os.path.join(settings.BASE_DIR, 'responsables.xlsx')
    wb = load_workbook(filename = filename)
    sh = wb['Responsables des clubs']
    current_club = None
    for i in range(3, 63):
        email = sh['I' + str(i)].value
        if (email==None):
            continue

        # club
        club_name = sh['A' + str(i)].value    
        if (club_name):
            club, created = Club.objects.get_or_create(name=club_name)
            current_club = club

        # user        
        try: 
            user = User.objects.get(email=email)
        except ObjectDoesNotExist:
            last_name = sh['C' + str(i)].value    
            first_name = sh['D' + str(i)].value 
            if (last_name and first_name):
                username = first_name.lower() + '.' + last_name.lower()
            else:
                username = email
            user = User(last_name=last_name, first_name=first_name, email=email, username=username)
            user.save()

        # contact
        try: 
            contact = Contact.objects.get(user_id=user.id)
        except ObjectDoesNotExist:
            address = sh['E' + str(i)].value
            npa = sh['F' + str(i)].value
            city = sh['G' + str(i)].value
            phone = sh['H' + str(i)].value
            contact = Contact(address=address, npa=npa, city=city, phone=phone, user_id=user.id)
            contact.save()

        # club responsable
        try: 
            club_responsable = ClubResponsable.objects.get(club_id=club.id, user_id=user.id)
        except ObjectDoesNotExist:
            function = sh['B' + str(i)].value
            club_responsable = ClubResponsable(club_id=club.id, user_id=user.id, function=function)
            club_responsable.save()

        return HttpResponse("Done")

def parse_teams(request):
    '''
    B(4, 10)  level_id 5
    B(12, 20) level_id 6
    E(12, 19) level_id 6
    B(22, 29) level_id 7
    E(22, 27) level_id 7
    '''
    filename = os.path.join(settings.BASE_DIR, 'Interclub 2020 - 2021.xlsx')
    wb = load_workbook(filename = filename)
    sh = wb['Equipes']

    for i in range(22, 27):
        txt = sh['E' + str(i)].value
        if (txt): 
            [reference, team_name] = txt.split(' ', 1)
            # team
            team, created = Team.objects.get_or_create(name=team_name)
            team.reference = reference
            team.level_id = 7
            team.save()


    return HttpResponse("Done")



        






        
        
        
        

