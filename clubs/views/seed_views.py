from django.conf import settings
from django.http import HttpResponse
from django.core.exceptions import ObjectDoesNotExist

from openpyxl import load_workbook
import os

from clubs.models import Club, ClubResponsible, Contact, Team, User


'''
order: 
import responsibles (users, clubs, club_responsible)
seed levels table
import teams
process teams
process users
'''
def import_responsibles(request): 
    filename = os.path.join(settings.BASE_DIR, 'excels/responsables.xlsx')
    wb = load_workbook(filename = filename)
    sh = wb['Responsables des clubs']
    current_club = None
    for i in range(3, 63):
        email = sh['I' + str(i)].value
        if (email==None):
            continue

        # club
        club_name = sh['A' + str(i)].value    
        if (club_name):
            club, created = Club.objects.get_or_create(name=club_name)
            current_club = club

        # user        
        try: 
            user = User.objects.get(email=email)
        except ObjectDoesNotExist:
            last_name = sh['C' + str(i)].value    
            first_name = sh['D' + str(i)].value 
            if (last_name and first_name):
                username = first_name.lower() + '.' + last_name.lower()
            else:
                username = email
            user = User(last_name=last_name, first_name=first_name, email=email, username=username)
            user.save()

        # contact
        try: 
            contact = Contact.objects.get(user_id=user.id)
        except ObjectDoesNotExist:
            address = sh['E' + str(i)].value
            npa = sh['F' + str(i)].value
            city = sh['G' + str(i)].value
            phone = sh['H' + str(i)].value
            contact = Contact(address=address, npa=npa, city=city, phone=phone, user_id=user.id)
            contact.save()

        # club responsable
        try: 
            club_responsable = ClubResponsible.objects.get(club_id=club.id, user_id=user.id)
        except ObjectDoesNotExist:
            function = sh['B' + str(i)].value
            club_responsable = ClubResponsible(club_id=club.id, user_id=user.id, function=function)
            club_responsable.save()

    return HttpResponse("Done")


def import_teams(request):
    '''
    get teams from excel 2020-2021
    B(4, 10)  level_id 5
    B(12, 20) level_id 6
    E(12, 19) level_id 6
    B(22, 29) level_id 7
    E(22, 27) level_id 7

    from excel 2021-2022
    '''
    filename = os.path.join(settings.BASE_DIR, 'excels/Inteclub version online 21-22.xlsx')
    wb = load_workbook(filename = filename)
    sh = wb['EQUIPES']

    groups= [
        {
            'first_column': 'B', 
            'start':4,
            'stop': 11,
            'level_id': 5
        }, 
        {
            'first_column': 'B', 
            'start':13,
            'stop': 19,
            'level_id': 6
        }, 
        {
            'first_column': 'E', 
            'start':13,
            'stop': 20,
            'level_id': 6
        }, 
        {
            'first_column': 'B', 
            'start':22,
            'stop': 29,
            'level_id': 7
        }, 
        {
            'first_column': 'E', 
            'start':22,
            'stop': 28,
            'level_id': 7
        }, 
    ]

    for group in groups:
        for i in range(group['start'], group['stop']):
            txt = sh[group['first_column'] + str(i)].value
            if (txt): 
                # [reference, team_name] = txt.split(' ', 1)
                team_name = txt
                # team
                team, created = Team.objects.get_or_create(name=team_name)
                # team.reference = reference
                team.level_id = group['level_id']
                team.save()


    return HttpResponse("Done")

def process_teams(request):
    '''
    try to add club_id by analysing names
    '''
    teams = Team.objects.filter(club_id__isnull=True)
    for team in teams:
        team_name = team.name.split()[0]
        club = Club.objects.filter(name__icontains=team_name).first()
        if club: 
            team.club_id = club.id
            team.save()
    return HttpResponse("Done")




def process_users(request): 
    '''
    try to add club_id by analysing club_responsable table
    '''
    users = User.objects.filter(club_id__isnull=True)
    for user in users:
        club_responsable = ClubResponsible.objects.filter(user_id = user.id).first()
        if club_responsable: 
            user.club_id = club_responsable.club_id
            user.save()
    return HttpResponse("Done")